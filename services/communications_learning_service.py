import csv
import hashlib
import json
import re
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from core.app_context import context
from services.logging_service import LoggingService
from services.time_service import TimeService


logger = LoggingService.get_logger("content")


class CommunicationsLearningService:

    VERSION = "communications-learning-v1"
    MIN_PATTERN_SAMPLE = 3
    FIELD_ALIASES = {
        "platform": ("platform", "network", "channel"),
        "post_id": ("post_id", "platform_post_id", "id", "source_identifier"),
        "communication_id": ("communication_id", "mfr_communication_id"),
        "package_id": ("package_id", "communication_package_id"),
        "media_package_id": ("media_package_id", "media_package"),
        "campaign": ("campaign", "campaign_name"),
        "program": ("program", "program_name"),
        "topic": ("topic", "topics", "tag"),
        "publication_date": ("publication_date", "published_at", "date", "post_date"),
        "publication_time": ("publication_time", "time", "post_time"),
        "reach": ("reach",),
        "impressions": ("impressions",),
        "views": ("views", "plays", "video_views"),
        "reactions": ("reactions", "likes", "like_count"),
        "comments": ("comments", "comment_count"),
        "shares": ("shares", "share_count"),
        "saves": ("saves", "save_count"),
        "link_clicks": ("link_clicks", "clicks"),
        "profile_visits": ("profile_visits",),
        "follower_gain": ("follower_gain", "new_followers"),
        "video_completion": ("video_completion", "completion_rate"),
        "watch_time": ("watch_time", "total_watch_time"),
        "average_watch_duration": ("average_watch_duration", "avg_watch_duration"),
        "caption": ("caption", "post_text", "text", "message"),
        "media_type": ("media_type", "format"),
        "media_ids": ("media_ids", "media", "assets"),
        "hashtags": ("hashtags",),
        "cta": ("cta", "call_to_action"),
        "boosted_post": ("boosted_post", "boosted", "paid")
    }

    def __init__(self, database=None):

        self.db = database or context.database
        self.last_metrics = {}

    ############################################################

    def preview_file(self, path, sample_size=25):

        started = time.perf_counter()
        path = Path(path)
        source_type = self._source_type(path)
        warnings = []

        if source_type == "unsupported":
            warnings.append(f"Unsupported learning source: {path.suffix}")
            return {
                "detected_format": source_type,
                "warnings": warnings,
                "confidence": 0
            }

        rows = list(self._iter_rows(path, source_type, limit=sample_size))
        fields = sorted({key for row in rows for key in row.keys()})
        mapping = self._auto_mapping(fields)
        normalized = [
            self._normalize_row(row, path, source_type, mapping, dry_run=True)
            for row in rows
        ]
        warnings.extend(
            warning
            for item in normalized
            for warning in item.get("warnings", [])
        )
        records = [item["record"] for item in normalized]

        return {
            "detected_format": source_type,
            "estimated_records": self._record_count(path, source_type),
            "mapped_fields": mapping,
            "detected_platforms": sorted({record["platform"] for record in records if record["platform"]}),
            "detected_topics": sorted({record["topic"] for record in records if record["topic"]}),
            "metrics_available": self._metric_availability(records),
            "linkable_records": sum(1 for record in records if record.get("communication_id") or record.get("post_id")),
            "invalid_rows": sum(1 for item in normalized if item.get("warnings")),
            "warnings": sorted(set(warnings))[:40],
            "confidence": self._preview_confidence(mapping, warnings),
            "sample_records": records[:sample_size],
            "inspection_seconds": round(time.perf_counter() - started, 3)
        }

    ############################################################

    def import_file(self, path, max_records=None, dry_run=False):

        started = time.perf_counter()
        path = Path(path)
        preview = self.preview_file(path)
        source_type = preview["detected_format"]

        if source_type == "unsupported":
            raise ValueError("Unsupported communications learning source.")

        summary = {
            "source_type": source_type,
            "source_file": str(path),
            "started_at": TimeService.utc_now_iso(),
            "records_processed": 0,
            "records_inserted": 0,
            "duplicates_skipped": 0,
            "invalid_records": 0,
            "warnings": [],
            "status": "dry_run" if dry_run else "running"
        }
        import_run_id = 0

        if not dry_run:
            import_run_id = self.db.create_communication_learning_import_run(summary)

        mapping = preview["mapped_fields"]

        for raw in self._iter_rows(path, source_type):
            if max_records and summary["records_inserted"] >= max_records:
                summary["status"] = "bounded_sample"
                break

            summary["records_processed"] += 1
            normalized = self._normalize_row(
                raw,
                path,
                source_type,
                mapping,
                import_run_id=import_run_id
            )
            summary["warnings"].extend(normalized["warnings"])
            record = normalized["record"]

            if not record["platform"] or not record["publication_date"]:
                summary["invalid_records"] += 1
                continue

            if dry_run:
                summary["records_inserted"] += 1
                continue

            result = self.db.save_communication_learning_record(record)

            if result["inserted"]:
                summary["records_inserted"] += 1
            else:
                summary["duplicates_skipped"] += 1

        summary["completed_at"] = TimeService.utc_now_iso()
        summary["duration_seconds"] = round(time.perf_counter() - started, 3)
        if summary["status"] == "running":
            summary["status"] = "completed"

        if not dry_run:
            self.db.update_communication_learning_import_run(
                import_run_id,
                summary
            )
            summary["import_run_id"] = import_run_id
            self.rebuild_summary()

        logger.info(
            "Communications learning import completed inserted=%s duplicates=%s elapsed=%s",
            summary["records_inserted"],
            summary["duplicates_skipped"],
            summary["duration_seconds"]
        )
        return summary

    def manual_entry(self, item):

        record = self._normalize_manual(item)
        result = self.db.save_communication_learning_record(record)
        self.rebuild_summary()
        return result

    ############################################################

    def rebuild_summary(self):

        records = self.records(limit=5000)
        summary = self._summary(records)
        confidence = summary["learning_confidence"]
        item = {
            "version": self.VERSION,
            "generated_at": TimeService.utc_now_iso(),
            "sample_count": len(records),
            "summary": summary,
            "confidence": confidence
        }
        self.db.save_communication_learning_summary(item)
        self.db.save_communication_learning_profile(
            {
                "profile_type": "performance",
                "profile_key": "mfr_communications",
                "version": self.VERSION,
                "generated_at": item["generated_at"],
                "sample_count": len(records),
                "confidence": confidence,
                "profile": summary,
                "source_summary": {
                    "record_count": len(records),
                    "benchmark_records_used": 0
                }
            }
        )
        return summary

    def dashboard(self):

        latest = self.db.latest_communication_learning_summary()

        if latest:
            return latest["summary"]

        return self.rebuild_summary()

    def records(self, filters=None, limit=500):

        return self.db.communication_learning_records(
            filters=filters or {},
            limit=limit
        )

    def recommendation_evidence(self, recommendation, limit=5):

        summary = self.dashboard()
        topic = self._token(
            recommendation.get("topic") or
            recommendation.get("category") or
            recommendation.get("editorial_angle")
        )
        topic_stats = summary.get("topics", {}).get(topic, {})
        sample_size = topic_stats.get("sample_size", 0)

        if sample_size < self.MIN_PATTERN_SAMPLE:
            return {
                "source": "mfr_historical_performance",
                "sample_size": sample_size,
                "confidence": self._confidence(sample_size),
                "recommendation": "Limited historical evidence.",
                "limitations": [
                    f"Needs at least {self.MIN_PATTERN_SAMPLE} similar MFR posts before drawing conclusions."
                ],
                "separate_from_benchmark": True
            }

        engagement = topic_stats.get("average_engagement_score", 0)
        label = "Strong" if engagement >= 70 else "Moderate" if engagement >= 45 else "Caution"
        return {
            "source": "mfr_historical_performance",
            "sample_size": sample_size,
            "confidence": self._confidence(sample_size),
            "average_normalized_engagement": engagement,
            "recommendation": label,
            "limitations": [
                "Correlation only; do not treat historical performance as causation.",
                "Benchmark inspiration is evaluated separately."
            ],
            "separate_from_benchmark": True
        }

    def performance_prediction(self, package):

        recommendation = package.get("source_recommendation") or package
        evidence = self.recommendation_evidence(recommendation)
        score = evidence.get("average_normalized_engagement", 0)
        return {
            "expected_engagement": (
                "high" if score >= 70 else "moderate" if score >= 45 else "unknown"
            ),
            "confidence": evidence.get("confidence", 0),
            "historical_support": evidence,
            "benchmark_support": package.get("benchmark_inspiration", []),
            "risks": evidence.get("limitations", []),
            "reasoning": (
                f"Based on {evidence.get('sample_size', 0)} similar MFR performance record(s)."
            ),
            "internal_only": True
        }

    ############################################################

    def _summary(self, records):

        active = [
            record
            for record in records
            if not record.get("exclude_from_learning")
        ]
        engagement_scores = [
            record["derived_metrics"].get("engagement_score", 0)
            for record in active
        ]
        by_topic = self._group_stats(active, "topic")
        by_platform = self._group_stats(active, "platform")
        by_weekday = self._weekday_stats(active)
        by_hour = self._hour_stats(active)
        media = self._group_stats(active, "media_type")
        fatigue = self._fatigue(active)

        return {
            "version": self.VERSION,
            "sample_count": len(active),
            "learning_confidence": self._confidence(len(active)),
            "baseline_engagement_score": round(self._avg(engagement_scores), 2),
            "top_performers": self._ranked(active, reverse=True)[:10],
            "underperformers": self._ranked(active, reverse=False)[:10],
            "topics": by_topic,
            "platforms": by_platform,
            "best_weekdays": by_weekday,
            "best_hours": by_hour,
            "media_performance": media,
            "reel_performance": self._reel_stats(active),
            "campaign_health": self._group_stats(active, "campaign"),
            "caption_intelligence": self._caption_stats(active),
            "fatigue": fatigue,
            "recent_successful_formats": self._successful_formats(active),
            "topics_cooling_down": fatigue.get("topic_fatigue", []),
            "topics_trending": self._trending_topics(active),
            "learning_limitations": self._limitations(len(active)),
            "benchmark_records_used": 0
        }

    def _group_stats(self, records, key):

        groups = defaultdict(list)
        for record in records:
            label = self._token(record.get(key) or "unknown")
            groups[label].append(record)

        result = {}
        for label, items in groups.items():
            scores = [
                item["derived_metrics"].get("engagement_score", 0)
                for item in items
            ]
            result[label] = {
                "sample_size": len(items),
                "average_engagement_score": round(self._avg(scores), 2),
                "confidence": self._confidence(len(items)),
                "limitations": self._limitations(len(items))
            }
        return result

    def _weekday_stats(self, records):

        groups = defaultdict(list)
        for record in records:
            try:
                dt = datetime.fromisoformat(record.get("publication_date", ""))
                label = dt.strftime("%A")
            except Exception:
                label = "unknown"
            groups[label].append(record)
        return self._simple_group(groups)

    def _hour_stats(self, records):

        groups = defaultdict(list)
        for record in records:
            text = str(record.get("publication_time") or "")
            label = text[:2] if len(text) >= 2 else "unknown"
            groups[label].append(record)
        return self._simple_group(groups)

    def _simple_group(self, groups):

        return {
            label: {
                "sample_size": len(items),
                "average_engagement_score": round(
                    self._avg([
                        item["derived_metrics"].get("engagement_score", 0)
                        for item in items
                    ]),
                    2
                ),
                "confidence": self._confidence(len(items))
            }
            for label, items in groups.items()
        }

    def _reel_stats(self, records):

        reels = [
            record
            for record in records
            if record.get("linked_context", {}).get("media_type") in ("reel", "video")
        ]
        return {
            "sample_size": len(reels),
            "average_engagement_score": round(
                self._avg([
                    record["derived_metrics"].get("engagement_score", 0)
                    for record in reels
                ]),
                2
            ),
            "confidence": self._confidence(len(reels)),
            "best_reel_duration": self._best_reel_duration(reels)
        }

    def _caption_stats(self, records):

        lengths = [
            record["derived_metrics"].get("caption_length", 0)
            for record in records
        ]
        emoji_density = [
            record["derived_metrics"].get("emoji_density", 0)
            for record in records
        ]
        return {
            "average_caption_length": round(self._avg(lengths), 2),
            "average_emoji_density": round(self._avg(emoji_density), 3),
            "question_posts": sum(1 for record in records if record["derived_metrics"].get("question_usage")),
            "common_ctas": Counter(
                record["derived_metrics"].get("cta_style", "none")
                for record in records
            ).most_common(6)
        }

    def _fatigue(self, records):

        topics = Counter(record.get("topic") or "unknown" for record in records)
        ctas = Counter(record["derived_metrics"].get("cta_style", "none") for record in records)
        hashtags = Counter(
            tag
            for record in records
            for tag in record["derived_metrics"].get("hashtags", [])
        )
        return {
            "topic_fatigue": self._fatigue_items(topics, "topic"),
            "cta_fatigue": self._fatigue_items(ctas, "cta"),
            "hashtag_fatigue": self._fatigue_items(hashtags, "hashtag"),
            "recommended_cooldown_days": 14 if records else 0
        }

    def _fatigue_items(self, counts, label):

        total = sum(counts.values()) or 1
        items = []
        for key, count in counts.most_common(8):
            share = count / total
            if count >= self.MIN_PATTERN_SAMPLE and share >= 0.35:
                items.append(
                    {
                        label: key,
                        "count": count,
                        "share": round(share, 2),
                        "recommendation": "cooldown"
                    }
                )
        return items

    def _successful_formats(self, records):

        strong = [
            record
            for record in records
            if record["derived_metrics"].get("engagement_score", 0) >= 70
        ]
        formats = Counter(
            record["derived_metrics"].get("media_package_type", "unknown")
            for record in strong
        )
        return [
            {
                "format": name,
                "count": count
            }
            for name, count in formats.most_common(6)
        ]

    def _trending_topics(self, records):

        ranked = self._group_stats(records, "topic")
        return [
            {
                "topic": topic,
                **stats
            }
            for topic, stats in sorted(
                ranked.items(),
                key=lambda item: item[1]["average_engagement_score"],
                reverse=True
            )[:6]
            if stats["sample_size"] >= 1
        ]

    def _ranked(self, records, reverse=True):

        return sorted(
            [
                {
                    "learning_id": record.get("learning_id"),
                    "platform": record.get("platform", ""),
                    "topic": record.get("topic", ""),
                    "campaign": record.get("campaign", ""),
                    "engagement_score": record["derived_metrics"].get("engagement_score", 0),
                    "publication_date": record.get("publication_date", "")
                }
                for record in records
            ],
            key=lambda item: item["engagement_score"],
            reverse=reverse
        )

    ############################################################

    def _normalize_row(self, raw, path, source_type, mapping, import_run_id=0, dry_run=False):

        def value(name):
            key = mapping.get(name)
            return raw.get(key, "") if key else ""

        metrics = {
            key: self._to_float(value(key))
            for key in (
                "reach",
                "impressions",
                "views",
                "reactions",
                "comments",
                "shares",
                "saves",
                "link_clicks",
                "profile_visits",
                "follower_gain",
                "video_completion",
                "watch_time",
                "average_watch_duration"
            )
        }
        platform = self._clean(value("platform")).lower()
        post_id = self._clean(value("post_id"))
        publication_date, publication_time, date_warning = self._date_parts(
            value("publication_date"),
            value("publication_time")
        )
        caption = self._clean(value("caption"))
        topic = self._token(value("topic") or self._infer_topic(caption))
        campaign = self._clean(value("campaign"))
        media_type = self._token(value("media_type") or "unknown")
        linked_media = self._ids(value("media_ids"))
        derived = self._derived_metrics(
            metrics,
            caption,
            media_type,
            value("cta"),
            value("hashtags")
        )
        communication_id = self._to_int(value("communication_id"))
        record = {
            "platform": platform,
            "post_id": post_id,
            "communication_id": communication_id,
            "package_id": self._clean(value("package_id")),
            "media_package_id": self._clean(value("media_package_id")),
            "campaign": campaign,
            "program": self._clean(value("program")),
            "topic": topic,
            "publication_date": publication_date,
            "publication_time": publication_time,
            "imported_from": str(path),
            "import_run_id": import_run_id,
            "metrics": metrics,
            "derived_metrics": derived,
            "linked_media": linked_media,
            "linked_context": {
                "media_type": media_type,
                "caption_excerpt": caption[:240],
                "hashtags": derived["hashtags"],
                "cta": value("cta"),
                "benchmark_inspiration": "not_used"
            },
            "source_type": source_type,
            "raw_record": dict(raw),
            "content_hash": self._hash(
                "|".join([
                    platform,
                    post_id,
                    str(communication_id),
                    publication_date,
                    caption
                ])
            ),
            "reviewed": False,
            "review_status": "unreviewed",
            "anomaly": False,
            "exclude_from_learning": False,
            "boosted_post": self._truthy(value("boosted_post")),
            "seasonal": False,
            "reviewer_notes": "",
            "created_at": TimeService.utc_now_iso(),
            "updated_at": TimeService.utc_now_iso()
        }
        warnings = []
        if date_warning:
            warnings.append(date_warning)
        if not platform:
            warnings.append("Missing platform.")
        if not publication_date:
            warnings.append("Missing publication date.")
        if not any(metrics.values()):
            warnings.append("No engagement metrics supplied.")
        return {
            "record": record,
            "warnings": warnings
        }

    def _normalize_manual(self, item):

        mapping = {
            key: key
            for key in item.keys()
        }
        return self._normalize_row(
            item,
            item.get("imported_from", "manual"),
            "manual",
            mapping
        )["record"]

    def _derived_metrics(self, metrics, caption, media_type, cta, hashtags):

        reach_base = metrics.get("reach") or metrics.get("impressions") or metrics.get("views") or 0
        interactions = (
            metrics.get("reactions", 0) +
            metrics.get("comments", 0) * 2 +
            metrics.get("shares", 0) * 3 +
            metrics.get("saves", 0) * 2 +
            metrics.get("link_clicks", 0) * 2
        )
        engagement_score = (
            min(100, (interactions / reach_base) * 1000)
            if reach_base
            else min(100, interactions)
        )
        words = re.findall(r"[A-Za-z0-9']+", caption or "")
        hashtag_list = self._hashtags(hashtags or caption)
        emoji_count = len(re.findall(r"[\U0001f300-\U0001faff\u2600-\u27bf]", caption or ""))
        return {
            "engagement_score": round(engagement_score, 2),
            "normalized_engagement": round(engagement_score, 2),
            "post_longevity": "unknown",
            "early_engagement": "unknown",
            "viral_acceleration": "unknown",
            "audience_interest": "high" if engagement_score >= 70 else "moderate" if engagement_score >= 35 else "unknown",
            "ctr": round((metrics.get("link_clicks", 0) / reach_base) * 100, 2) if reach_base else 0,
            "community_interaction": metrics.get("comments", 0) + metrics.get("shares", 0),
            "educational_success": self._success_score(caption, engagement_score, ("safety", "learn", "check", "prevent")),
            "recruitment_success": self._success_score(caption, engagement_score, ("recruit", "join", "serve", "volunteer")),
            "public_safety_success": self._success_score(caption, engagement_score, ("warning", "prevent", "alarm", "safety")),
            "campaign_success": engagement_score,
            "seasonality": self._season(caption),
            "evergreen_value": 75 if self._contains_any(caption.lower(), ("check", "practice", "prevent", "learn")) else 45,
            "repeatability": "repeatable" if engagement_score >= 50 else "watch",
            "confidence": self._confidence(1 if reach_base else 0),
            "limitations": self._metric_limitations(metrics),
            "caption_length": len(caption or ""),
            "paragraph_count": len([p for p in str(caption).split("\n") if p.strip()]) or 1,
            "emoji_density": round(emoji_count / max(1, len(words)), 3),
            "question_usage": "?" in str(caption),
            "hook_effectiveness": "unknown_without_ab_test",
            "cta_style": self._cta_style(cta or caption),
            "reading_difficulty": "plain" if len(words) < 90 else "moderate",
            "estimated_attention": "strong" if engagement_score >= 70 else "standard",
            "hashtags": hashtag_list,
            "media_package_type": media_type
        }

    ############################################################

    def _iter_rows(self, path, source_type, limit=None):

        path = Path(path)
        if source_type == "csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                for index, row in enumerate(reader):
                    if limit and index >= limit:
                        break
                    yield dict(row)
        elif source_type == "json":
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                data = data.get("posts") or data.get("records") or data.get("items") or [data]
            for index, row in enumerate(data):
                if limit and index >= limit:
                    break
                if isinstance(row, dict):
                    yield row
        elif source_type == "xlsx":
            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows)]
            for index, values in enumerate(rows):
                if limit and index >= limit:
                    break
                yield {
                    headers[column]: values[column]
                    for column in range(min(len(headers), len(values)))
                }

    def _source_type(self, path):

        suffix = Path(path).suffix.lower()
        if suffix == ".csv":
            return "csv"
        if suffix == ".json":
            return "json"
        if suffix == ".xlsx":
            return "xlsx"
        return "unsupported"

    def _record_count(self, path, source_type):

        try:
            return sum(1 for _ in self._iter_rows(path, source_type))
        except Exception:
            return 0

    def _auto_mapping(self, fields):

        lowered = {str(field).lower().strip(): field for field in fields}
        mapping = {}
        for canonical, aliases in self.FIELD_ALIASES.items():
            for alias in aliases:
                if alias in lowered:
                    mapping[canonical] = lowered[alias]
                    break
        return mapping

    def _metric_availability(self, records):

        names = Counter()
        for record in records:
            for key, value in record.get("metrics", {}).items():
                if value:
                    names[key] += 1
        return dict(names)

    def _preview_confidence(self, mapping, warnings):

        required = ("platform", "publication_date")
        confidence = sum(35 for key in required if mapping.get(key))
        if any(mapping.get(key) for key in ("reach", "impressions", "views", "reactions")):
            confidence += 25
        return max(0, min(100, confidence - min(30, len(warnings) * 3)))

    def _date_parts(self, date_value, time_value):

        text = self._clean(date_value)
        time_text = self._clean(time_value)
        if not text:
            return "", time_text, "Missing publication date."
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except Exception:
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
                try:
                    parsed = datetime.strptime(text, fmt)
                    break
                except Exception:
                    parsed = None
            if parsed is None:
                return "", time_text, f"Ambiguous publication date: {text}"
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        if not time_text:
            time_text = parsed.strftime("%H:%M")
        return parsed.date().isoformat(), time_text, ""

    def _best_reel_duration(self, records):

        durations = [
            record.get("metrics", {}).get("average_watch_duration", 0)
            for record in records
            if record.get("metrics", {}).get("average_watch_duration", 0)
        ]
        return round(self._avg(durations), 2) if durations else 0

    def _metric_limitations(self, metrics):

        limitations = ["Correlation only; no causation implied."]
        if not (metrics.get("reach") or metrics.get("impressions") or metrics.get("views")):
            limitations.append("No reach, impressions, or views denominator supplied.")
        if metrics.get("boosted_post"):
            limitations.append("Boosted or paid distribution may distort performance.")
        return limitations

    def _success_score(self, caption, engagement_score, terms):

        return engagement_score if self._contains_any(str(caption).lower(), terms) else 0

    def _cta_style(self, text):

        lower = str(text or "").lower()
        if self._contains_any(lower, ("apply", "join")):
            return "recruitment"
        if self._contains_any(lower, ("learn", "visit", "read")):
            return "learn_more"
        if self._contains_any(lower, ("share", "comment", "tag")):
            return "engagement"
        if self._contains_any(lower, ("check", "test", "practice")):
            return "safety_action"
        return "none"

    def _infer_topic(self, text):

        lower = str(text or "").lower()
        for topic, terms in {
            "recruitment": ("recruit", "join", "volunteer"),
            "training": ("training", "drill", "evolution"),
            "public_education": ("safety", "alarm", "prevent"),
            "community": ("community", "event", "families"),
            "incident": ("incident", "scene", "responded")
        }.items():
            if self._contains_any(lower, terms):
                return topic
        return "general"

    def _season(self, text):

        lower = str(text or "").lower()
        if self._contains_any(lower, ("winter", "ice", "heating")):
            return "winter"
        if self._contains_any(lower, ("summer", "heat", "water")):
            return "summer"
        if self._contains_any(lower, ("spring", "wildfire", "flood")):
            return "spring"
        if self._contains_any(lower, ("fall", "school", "smoke alarm")):
            return "fall"
        return "evergreen"

    def _confidence(self, sample_size):

        return min(95, max(10, int(sample_size) * 12))

    def _limitations(self, sample_size):

        if sample_size < self.MIN_PATTERN_SAMPLE:
            return [f"Only {sample_size} sample(s); treat as directional."]
        return ["Evidence is historical MFR performance, not a guarantee."]

    def _avg(self, values):

        values = [float(value or 0) for value in values]
        return sum(values) / len(values) if values else 0

    def _ids(self, value):

        if isinstance(value, list):
            return [self._to_int(item) for item in value if self._to_int(item)]
        return [
            self._to_int(part)
            for part in re.split(r"[,; ]+", str(value or ""))
            if self._to_int(part)
        ]

    def _hashtags(self, text):

        return list(dict.fromkeys(re.findall(r"#[A-Za-z0-9_]+", str(text or ""))))[:12]

    def _hash(self, text):

        return hashlib.sha256(str(text or "").encode("utf-8")).hexdigest()

    def _token(self, value):

        token = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower())
        return token.strip("_") or "general"

    def _clean(self, value):

        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _contains_any(self, lower, terms):

        return any(term in lower for term in terms)

    def _truthy(self, value):

        return str(value or "").strip().lower() in ("1", "true", "yes", "y")

    def _to_float(self, value):

        try:
            return float(value or 0)
        except Exception:
            return 0

    def _to_int(self, value):

        try:
            return int(float(value or 0))
        except Exception:
            return 0
