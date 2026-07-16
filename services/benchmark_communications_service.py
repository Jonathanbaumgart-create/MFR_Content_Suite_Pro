import csv
import hashlib
import json
import re
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from core.app_context import context
from services.logging_service import LoggingService
from services.time_service import TimeService


logger = LoggingService.get_logger("content")


class BenchmarkCommunicationsService:

    PREVIEW_LIMIT = 25
    BATCH_SIZE = 100
    SOURCE_DEPARTMENTS = (
        "South Metro Fire Rescue",
        "Central Pierce Fire & Rescue",
        "Orange County Fire Authority",
        "North Metro Fire Rescue District",
        "Calgary Fire Department",
        "Edmonton Fire Rescue Services",
        "Mesa Fire & Medical Department",
        "Puget Sound Fire"
    )
    FIELD_ALIASES = {
        "department": ("department", "source_department", "agency", "page", "organization"),
        "platform": ("platform", "network", "channel", "source_platform"),
        "publication_date": ("publication_date", "published_at", "date", "post_date", "created_at", "timestamp"),
        "post_text": ("post_text", "text", "caption", "body", "message", "description", "content"),
        "headline": ("headline", "title", "subject"),
        "source_url": ("source_url", "url", "permalink", "link", "href"),
        "post_id": ("post_id", "id", "source_identifier", "platform_post_id"),
        "media_type": ("media_type", "format", "asset_type"),
        "photo_count": ("photo_count", "photos", "image_count"),
        "video_count": ("video_count", "videos", "video_count"),
        "reel": ("reel", "is_reel", "short_video", "short_video_indicator"),
        "duration": ("duration", "duration_seconds", "video_duration"),
        "reactions": ("reactions", "likes", "like_count", "reaction_count"),
        "comments": ("comments", "comment_count", "comments_count"),
        "shares": ("shares", "share_count", "shares_count"),
        "views": ("views", "view_count", "plays", "play_count"),
        "saves": ("saves", "save_count"),
        "hashtags": ("hashtags", "hash_tags"),
        "cta": ("cta", "call_to_action"),
        "campaign": ("campaign", "campaign_name"),
        "topic": ("topic", "topics", "tags", "labels"),
        "audience": ("audience", "target_audience"),
        "editorial_angle": ("editorial_angle", "angle", "story_angle"),
        "notes": ("notes", "comments_internal"),
        "source_filename": ("source_filename", "source_file", "filename")
    }
    TOPIC_TERMS = {
        "recruitment": ("recruit", "hiring", "join", "volunteer", "career"),
        "training": ("training", "drill", "academy", "evolution", "practice"),
        "fire_prevention": ("fire prevention", "escape plan", "cooking safety", "prevention"),
        "smoke_alarm": ("smoke alarm", "smoke alarms", "carbon monoxide"),
        "community_event": ("community", "event", "open house", "parade", "visit"),
        "incident_information": ("incident", "crews", "responded", "scene", "update"),
        "apparatus": ("engine", "ladder", "rescue", "apparatus", "truck"),
        "wildfire": ("wildfire", "brush", "grass fire", "vegetation"),
        "water_safety": ("water safety", "river", "lake", "swim"),
        "recognition": ("thank", "congratulations", "recognize", "award")
    }
    CTA_TERMS = {
        "safety_action": ("check", "test", "replace", "practice", "make sure"),
        "learn_more": ("learn more", "visit", "read more"),
        "recruitment": ("apply", "join", "serve"),
        "engagement": ("share", "comment", "tell us", "tag"),
        "event": ("register", "attend", "stop by")
    }

    def __init__(self, database=None):

        self.db = database or context.database
        self.last_metrics = {}

    ############################################################

    def preview_file(self, path, mapping=None, sample_size=PREVIEW_LIMIT):

        started = time.perf_counter()
        path = Path(path)
        source_type = self._source_type(path)
        warnings = []

        if source_type == "unsupported":
            warnings.append(f"Unsupported benchmark file type: {path.suffix}")
            return self._preview_result(
                path,
                source_type,
                [],
                {},
                warnings,
                started
            )

        rows = list(self._iter_rows(path, source_type, limit=sample_size + 1))
        fields = sorted({key for row in rows for key in row.keys()})
        mapping = mapping or self._auto_mapping(fields)
        normalized = [
            self._normalize_row(row, path, source_type, mapping, dry_run=True)
            for row in rows[:sample_size]
        ]
        warnings.extend(
            warning
            for item in normalized
            for warning in item.get("warnings", [])
        )

        return self._preview_result(
            path,
            source_type,
            normalized,
            mapping,
            warnings,
            started,
            fields=fields,
            record_count=self._record_count(path, source_type)
        )

    ############################################################

    def import_file(
        self,
        path,
        mapping=None,
        max_records=None,
        dry_run=False,
        progress_callback=None,
        cancel_check=None
    ):

        started = time.perf_counter()
        path = Path(path)
        preview = self.preview_file(path, mapping=mapping)
        source_type = preview["detected_format"]

        if source_type == "unsupported":
            raise ValueError("Unsupported benchmark import source.")

        summary = {
            "source_type": source_type,
            "source_file": str(path),
            "started_at": TimeService.utc_now_iso(),
            "records_processed": 0,
            "records_inserted": 0,
            "duplicates_skipped": 0,
            "invalid_records": 0,
            "warnings": [],
            "patterns_generated": 0,
            "status": "dry_run" if dry_run else "running"
        }
        import_run_id = 0

        if not dry_run:
            import_run_id = self.db.create_benchmark_import_run(summary)

        inserted_records = []
        mapping = mapping or preview["mapped_fields"]

        for raw in self._iter_rows(path, source_type):
            if cancel_check and cancel_check():
                summary["status"] = "cancelled"
                break

            if max_records and summary["records_inserted"] >= max_records:
                summary["status"] = "bounded_sample"
                break

            summary["records_processed"] += 1
            normalized = self._normalize_row(
                raw,
                path,
                source_type,
                mapping,
                import_run_id=import_run_id
            )
            summary["warnings"].extend(normalized["warnings"])
            record = normalized["record"]

            if not record["source_department"] or not record["original_text"]:
                summary["invalid_records"] += 1
                continue

            if dry_run:
                summary["records_inserted"] += 1
                inserted_records.append(record)
                continue

            department_id = self.db.save_benchmark_department(
                {
                    "name": record["source_department"],
                    "created_at": TimeService.utc_now_iso(),
                    "updated_at": TimeService.utc_now_iso()
                }
            )
            record["department_id"] = department_id
            result = self.db.save_benchmark_record(record)

            if result["inserted"]:
                summary["records_inserted"] += 1
                record["benchmark_id"] = result["benchmark_id"]
                inserted_records.append(record)
            else:
                summary["duplicates_skipped"] += 1

            if progress_callback and summary["records_processed"] % self.BATCH_SIZE == 0:
                progress_callback(dict(summary))

        if not dry_run:
            patterns = self.discover_patterns(records=inserted_records)
            for pattern in patterns:
                self.db.save_benchmark_pattern(pattern)
            summary["patterns_generated"] = len(patterns)

        summary["completed_at"] = TimeService.utc_now_iso()
        summary["duration_seconds"] = round(time.perf_counter() - started, 3)
        if summary["status"] == "running":
            summary["status"] = "completed"

        if not dry_run:
            self.db.update_benchmark_import_run(import_run_id, summary)
            summary["import_run_id"] = import_run_id

        self.last_metrics = {
            "total_seconds": summary["duration_seconds"],
            "records_processed": summary["records_processed"],
            "records_inserted": summary["records_inserted"],
            "dry_run": dry_run
        }
        logger.info(
            "Benchmark import completed source=%s inserted=%s duplicates=%s elapsed=%s",
            path.name,
            summary["records_inserted"],
            summary["duplicates_skipped"],
            summary["duration_seconds"]
        )
        return summary

    ############################################################

    def search(self, filters=None, limit=100, offset=0):

        return self.db.benchmark_records(
            filters=filters or {},
            limit=limit,
            offset=offset
        )

    def insights(self):

        return self.db.benchmark_insights()

    def advisory_patterns(self, recommendation=None, limit=5):

        recommendation = recommendation or {}
        filters = {}
        topic = recommendation.get("topic") or recommendation.get("category", "")

        if topic:
            filters["topic"] = self._token(topic)

        patterns = self.db.benchmark_patterns(filters=filters, limit=limit)

        if not patterns:
            patterns = self.db.benchmark_patterns(
                filters={"applicability": "Highly applicable"},
                limit=limit
            )

        return [
            {
                "source": "external_benchmark",
                "source_department": pattern.get("source_department", ""),
                "observed_format": pattern.get("title", ""),
                "why_it_may_work": pattern.get("description", ""),
                "how_to_adapt_for_mfr": pattern.get("adaptation_notes", ""),
                "source_limitations": pattern.get("limitations", ""),
                "applicability": pattern.get("applicability", ""),
                "evidence_count": pattern.get("evidence_count", 0),
                "internal_only": True
            }
            for pattern in patterns
        ]

    ############################################################

    def discover_patterns(self, records=None, limit=25):

        records = records if records is not None else self.search(limit=500)
        groups = defaultdict(list)

        for record in records:
            analysis = record.get("normalized_analysis") or {}
            key = (
                record.get("source_department", ""),
                record.get("source_platform", ""),
                analysis.get("format", record.get("editorial_angle", "")),
                record.get("topic", ""),
                record.get("media_type", "")
            )
            groups[key].append(record)

        patterns = []
        now = TimeService.utc_now_iso()

        for key, items in groups.items():
            department, platform, format_name, topic, media_type = key
            evidence_count = len(items)

            if evidence_count < 1:
                continue

            exemplar = items[0]
            analysis = exemplar.get("normalized_analysis") or {}
            engagement_supported = any(
                item.get("engagement_available")
                for item in items
            )
            applicability = self._applicability(exemplar, analysis)
            patterns.append(
                {
                    "pattern_key": self._hash("|".join(str(part) for part in key)),
                    "pattern_type": analysis.get("format", format_name) or "general",
                    "title": self._pattern_title(analysis, media_type),
                    "description": self._pattern_description(analysis),
                    "source_department": department,
                    "source_platform": platform,
                    "topic": topic,
                    "editorial_angle": exemplar.get("editorial_angle", ""),
                    "media_type": media_type,
                    "reel_pattern": bool(exemplar.get("reel_indicator")),
                    "evidence_count": evidence_count,
                    "benchmark_ids": [
                        item.get("benchmark_id", 0)
                        for item in items
                        if item.get("benchmark_id")
                    ],
                    "engagement_basis": (
                        "Engagement supplied for at least one supporting record."
                        if engagement_supported
                        else "No comparable engagement evidence supplied."
                    ),
                    "applicability": applicability["label"],
                    "applicability_reason": applicability["reason"],
                    "adaptation_notes": applicability["adaptation"],
                    "limitations": self._limitations(items),
                    "human_status": "unreviewed",
                    "created_at": now,
                    "updated_at": now
                }
            )

        patterns.sort(
            key=lambda item: (
                item["evidence_count"],
                item["applicability"] == "Highly applicable"
            ),
            reverse=True
        )
        return patterns[:limit]

    ############################################################

    def _normalize_row(
        self,
        raw,
        path,
        source_type,
        mapping,
        import_run_id=0,
        dry_run=False
    ):

        warnings = []

        def value(name):
            key = mapping.get(name)
            return self._clean(raw.get(key, "")) if key else ""

        text = value("post_text")
        headline = value("headline")
        department = value("department")
        platform = value("platform").lower()
        source_date_text = value("publication_date")
        date_utc, date_warning = self._parse_date(source_date_text)

        if date_warning:
            warnings.append(date_warning)

        media_type = self._media_type(
            value("media_type"),
            value("photo_count"),
            value("video_count"),
            value("reel")
        )
        engagement = self._engagement(raw, mapping)
        analysis = self._analyze_text(
            headline,
            text,
            media_type,
            value("reel"),
            engagement
        )
        topic = value("topic") or analysis["topic"]
        campaign = value("campaign") or analysis["campaign"]
        editorial_angle = value("editorial_angle") or analysis["editorial_angle"]
        applicability = self._applicability(
            {
                "source_department": department,
                "media_type": media_type,
                "topic": topic,
                "campaign": campaign,
                "editorial_angle": editorial_angle
            },
            analysis
        )
        content_hash = self._hash(
            "|".join(
                [
                    department.lower(),
                    platform,
                    source_date_text,
                    text.lower(),
                    value("post_id"),
                    value("source_url")
                ]
            )
        )
        record = {
            "department_id": 0,
            "source_department": department,
            "source_platform": platform or "unknown",
            "source_date_text": source_date_text,
            "source_date_utc": date_utc,
            "source_url": value("source_url"),
            "source_identifier": value("post_id"),
            "source_file": str(path),
            "import_run_id": import_run_id,
            "headline": headline,
            "original_text": text,
            "normalized_analysis": analysis,
            "media_type": media_type,
            "photo_count": self._to_int(value("photo_count")),
            "video_count": self._to_int(value("video_count")),
            "reel_indicator": analysis["reel_indicator"],
            "duration_seconds": self._to_float(value("duration")),
            "raw_engagement": engagement["raw"],
            "engagement_available": engagement["available"],
            "engagement_status": engagement["status"],
            "engagement_indicator": engagement["indicator"],
            "hashtags": self._hashtags(value("hashtags") or text),
            "cta": value("cta") or analysis["cta_type"],
            "campaign": campaign,
            "topic": self._token(topic),
            "audience": value("audience") or analysis["audience"],
            "editorial_angle": editorial_angle,
            "raw_metadata": {
                "source_type": source_type,
                "source_filename": value("source_filename") or Path(path).name,
                "notes": value("notes"),
                "raw_keys": sorted(raw.keys())
            },
            "reviewed": False,
            "review_status": "unreviewed",
            "applicability": applicability["label"],
            "copyright_status": (
                "source_attribution_preserved"
                if value("source_url") or value("post_id")
                else "missing_source_reference"
            ),
            "content_hash": content_hash,
            "imported_at": TimeService.utc_now_iso()
        }

        if not department:
            warnings.append("Missing source department.")

        if not text:
            warnings.append("Missing post text.")

        if record["copyright_status"] == "missing_source_reference":
            warnings.append("Missing source URL or source identifier.")

        return {
            "record": record,
            "warnings": warnings
        }

    ############################################################

    def _analyze_text(self, headline, text, media_type, reel_value, engagement):

        combined = self._clean(" ".join([headline, text]))
        lower = combined.lower()
        topic = self._topic(lower)
        cta_type = self._cta_type(lower)
        hook_style = self._hook_style(combined)
        opening_style = self._opening_style(combined)
        emoji_count = len(re.findall(r"[\U0001f300-\U0001faff\u2600-\u27bf]", combined))
        hashtags = self._hashtags(combined)
        paragraph_count = len([p for p in combined.split("\n") if p.strip()]) or 1
        words = re.findall(r"[A-Za-z0-9']+", combined)
        reel_indicator = self._truthy(reel_value) or media_type in ("reel", "short_video")
        format_name = self._format(media_type, reel_indicator, lower)
        structure = self._educational_structure(lower)

        return {
            "opening_style": opening_style,
            "hook_style": hook_style,
            "educational_structure": structure,
            "emotional_framing": self._contains_any(lower, ("proud", "thank", "honour", "heart")),
            "community_framing": self._contains_any(lower, ("community", "neighbour", "families", "residents")),
            "operational_framing": self._contains_any(lower, ("crews", "incident", "responded", "training")),
            "recruitment_framing": self._contains_any(lower, ("recruit", "join", "career", "volunteer")),
            "public_safety_framing": self._contains_any(lower, ("safety", "prevent", "check", "warning")),
            "cta_type": cta_type,
            "hashtag_strategy": "focused" if 1 <= len(hashtags) <= 5 else "none" if not hashtags else "heavy",
            "paragraph_count": paragraph_count,
            "caption_length": len(combined),
            "word_count": len(words),
            "emoji_count": emoji_count,
            "question_use": "?" in combined,
            "urgency_style": "urgent" if self._contains_any(lower, ("now", "today", "warning", "immediately")) else "standard",
            "storytelling_arc": "setup_action_takeaway" if structure != "single_message" else "single_message",
            "format": format_name,
            "reel_indicator": reel_indicator,
            "reel_concept": self._reel_concept(lower, media_type, reel_indicator),
            "video_hook": self._video_hook(lower, reel_indicator),
            "campaign": self._campaign(lower),
            "topic": topic,
            "audience": self._audience(lower, topic),
            "editorial_angle": self._editorial_angle(topic, format_name),
            "seasonality": self._seasonality(lower),
            "evergreen_value": self._evergreen(lower),
            "reuse_potential": self._reuse_potential(lower, engagement),
            "engagement": engagement
        }

    ############################################################

    def _preview_result(
        self,
        path,
        source_type,
        normalized,
        mapping,
        warnings,
        started,
        fields=None,
        record_count=0
    ):

        records = [
            item["record"]
            for item in normalized
        ]
        return {
            "detected_format": source_type,
            "detected_departments": sorted({
                record.get("source_department", "")
                for record in records
                if record.get("source_department")
            }),
            "detected_platforms": sorted({
                record.get("source_platform", "")
                for record in records
                if record.get("source_platform")
            }),
            "estimated_records": record_count or len(records),
            "detected_fields": fields or [],
            "mapped_fields": mapping,
            "date_formats": sorted({
                "parsed" if record.get("source_date_utc") else "missing_or_ambiguous"
                for record in records
            }),
            "engagement_availability": {
                "with_engagement": sum(1 for record in records if record.get("engagement_available")),
                "missing_engagement": sum(1 for record in records if not record.get("engagement_available"))
            },
            "media_type_coverage": Counter(record.get("media_type", "unknown") for record in records),
            "duplicates": self._preview_duplicates(records),
            "invalid_rows": sum(
                1
                for item in normalized
                if item["warnings"]
            ),
            "missing_source_attribution": sum(
                1
                for record in records
                if record.get("copyright_status") == "missing_source_reference"
            ),
            "warnings": sorted(set(warnings))[:50],
            "confidence": self._preview_confidence(source_type, mapping, warnings),
            "sample_records": records[: self.PREVIEW_LIMIT],
            "inspection_seconds": round(time.perf_counter() - started, 3)
        }

    ############################################################

    def _iter_rows(self, path, source_type, limit=None):

        path = Path(path)

        if source_type == "csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                for index, row in enumerate(reader):
                    if limit and index >= limit:
                        break
                    yield dict(row)

        elif source_type == "json":
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                data = data.get("posts") or data.get("records") or data.get("items") or [data]
            for index, row in enumerate(data):
                if limit and index >= limit:
                    break
                if isinstance(row, dict):
                    yield row

        elif source_type == "xlsx":
            try:
                from openpyxl import load_workbook
            except Exception as ex:
                raise ValueError("XLSX support requires openpyxl.") from ex

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [
                str(value or "").strip()
                for value in next(rows)
            ]
            for index, values in enumerate(rows):
                if limit and index >= limit:
                    break
                yield {
                    headers[column]: values[column]
                    for column in range(min(len(headers), len(values)))
                }

    def _source_type(self, path):

        suffix = Path(path).suffix.lower()

        if suffix == ".csv":
            return "csv"

        if suffix == ".json":
            return "json"

        if suffix == ".xlsx":
            return "xlsx"

        return "unsupported"

    def _record_count(self, path, source_type):

        try:
            return sum(1 for _ in self._iter_rows(path, source_type))
        except Exception:
            return 0

    def _auto_mapping(self, fields):

        lowered = {
            field.lower().strip(): field
            for field in fields
        }
        mapping = {}

        for canonical, aliases in self.FIELD_ALIASES.items():
            for alias in aliases:
                if alias in lowered:
                    mapping[canonical] = lowered[alias]
                    break

        return mapping

    ############################################################

    def _engagement(self, raw, mapping):

        metrics = {}

        for key in ("reactions", "comments", "shares", "views", "saves"):
            field = mapping.get(key)
            if not field:
                continue
            metrics[key] = self._to_float(raw.get(field))

        available = any(value > 0 for value in metrics.values())
        views = metrics.get("views", 0)
        interactions = (
            metrics.get("reactions", 0) +
            metrics.get("comments", 0) * 2 +
            metrics.get("shares", 0) * 3 +
            metrics.get("saves", 0) * 2
        )

        if not available:
            return {
                "raw": metrics,
                "available": False,
                "status": "missing_engagement",
                "indicator": 0,
                "limitations": ["No engagement metrics supplied."]
            }

        if views:
            indicator = min(100, (interactions / max(1, views)) * 1000)
            status = "comparable_within_platform_views"
        else:
            indicator = min(100, interactions)
            status = "not_cross_platform_comparable"

        return {
            "raw": metrics,
            "available": True,
            "status": status,
            "indicator": round(indicator, 2),
            "limitations": [
                "Compare only within similar platforms and source context.",
                "Follower counts are not inferred."
            ]
        }

    def _applicability(self, record, analysis):

        topic = str(record.get("topic") or analysis.get("topic") or "").lower()
        media_type = str(record.get("media_type") or "").lower()
        text = " ".join([topic, str(record.get("campaign", "")), str(record.get("editorial_angle", ""))]).lower()

        if self._contains_any(text, ("recruitment", "training", "fire_prevention", "smoke_alarm", "community_event")):
            return {
                "label": "Highly applicable",
                "reason": "The pattern aligns with paid-on-call municipal communications and common MFR story types.",
                "adaptation": "Adapt the structure using MFR media, Morden context, and Jonathan-approved department voice."
            }

        if media_type in ("reel", "short_video", "video"):
            return {
                "label": "Applicable with adaptation",
                "reason": "Video-first pattern may work for MFR if footage is available and staff time is realistic.",
                "adaptation": "Use shorter clips, avoid copying branded series, and keep captions MFR-specific."
            }

        if self._contains_any(text, ("large_agency", "daily staffing", "metro")):
            return {
                "label": "Limited applicability",
                "reason": "The pattern may assume staffing or production capacity beyond MFR's normal workflow.",
                "adaptation": "Simplify to one strong post or a small carousel."
            }

        return {
            "label": "Applicable with adaptation",
            "reason": "The format has general communications value but needs local context and voice adaptation.",
            "adaptation": "Use the lesson as inspiration only; do not copy wording or source branding."
        }

    ############################################################

    def _topic(self, lower):

        for topic, terms in self.TOPIC_TERMS.items():
            if self._contains_any(lower, terms):
                return topic

        return "general"

    def _cta_type(self, lower):

        for cta_type, terms in self.CTA_TERMS.items():
            if self._contains_any(lower, terms):
                return cta_type

        return "none"

    def _hook_style(self, text):

        stripped = text.strip()

        if not stripped:
            return "none"

        if stripped.endswith("?") or "?" in stripped[:90]:
            return "question_hook"

        if re.search(r"\b(today|now|warning|alert)\b", stripped[:120], re.I):
            return "timely_alert"

        if re.search(r"\bthank|proud|congrat", stripped[:120], re.I):
            return "recognition_hook"

        return "direct_context"

    def _opening_style(self, text):

        first = text.strip().split(".")[0][:120].lower()

        if "did you know" in first:
            return "educational_question"

        if first.startswith(("thank", "congrat")):
            return "recognition"

        if len(first) < 55:
            return "short_hook"

        return "context_lead"

    def _format(self, media_type, reel_indicator, lower):

        if reel_indicator:
            return "reel_or_short_video"

        if media_type == "carousel" or "carousel" in lower:
            return "carousel"

        if media_type == "video":
            return "video_post"

        if media_type == "photo":
            return "photo_post"

        if self._contains_any(lower, ("step 1", "three", "first", "second", "third")):
            return "educational_sequence"

        return "standard_caption"

    def _educational_structure(self, lower):

        if self._contains_any(lower, ("first", "second", "third", "three things", "steps")):
            return "step_by_step"

        if self._contains_any(lower, ("why", "because", "helps")):
            return "explain_why"

        return "single_message"

    def _media_type(self, value, photo_count, video_count, reel):

        lower = str(value or "").lower()

        if self._truthy(reel) or "reel" in lower or "short" in lower:
            return "reel"

        if "carousel" in lower:
            return "carousel"

        if "video" in lower or self._to_int(video_count) > 0:
            return "video"

        if "photo" in lower or "image" in lower or self._to_int(photo_count) > 0:
            return "photo"

        return "unknown"

    def _reel_concept(self, lower, media_type, reel_indicator):

        if not reel_indicator and media_type not in ("video", "reel"):
            return ""

        if self._contains_any(lower, ("training", "drill")):
            return "training_progression"

        if self._contains_any(lower, ("engine", "ladder", "apparatus")):
            return "apparatus_reveal"

        if self._contains_any(lower, ("community", "families", "kids")):
            return "community_interaction"

        return "short_action_sequence"

    def _video_hook(self, lower, reel_indicator):

        if not reel_indicator:
            return ""

        if self._contains_any(lower, ("before", "after")):
            return "before_after"

        if self._contains_any(lower, ("watch", "see")):
            return "visual_invitation"

        return "first_three_second_action"

    def _campaign(self, lower):

        if "fire prevention" in lower:
            return "Fire Prevention"

        if "recruit" in lower:
            return "Recruitment"

        if "smoke alarm" in lower:
            return "Smoke Alarm Safety"

        return ""

    def _audience(self, lower, topic):

        if self._contains_any(lower, ("kids", "children", "students", "school")):
            return "Families and students"

        if topic == "recruitment":
            return "Prospective firefighters"

        return "Community residents"

    def _editorial_angle(self, topic, format_name):

        if topic == "recruitment":
            return "Recruitment"

        if topic in ("training", "apparatus"):
            return "Training Highlight" if topic == "training" else "Apparatus Feature"

        if topic in ("fire_prevention", "smoke_alarm", "water_safety"):
            return "Public Education"

        if format_name == "reel_or_short_video":
            return "Reel Concept"

        return topic.replace("_", " ").title()

    def _seasonality(self, lower):

        seasons = []
        for season, terms in {
            "winter": ("winter", "ice", "holiday", "heating"),
            "spring": ("spring", "flood", "wildfire"),
            "summer": ("summer", "heat", "water"),
            "fall": ("fall", "school", "smoke alarm")
        }.items():
            if self._contains_any(lower, terms):
                seasons.append(season)
        return seasons

    def _evergreen(self, lower):

        return 80 if self._contains_any(lower, ("check", "practice", "prevent", "learn")) else 45

    def _reuse_potential(self, lower, engagement):

        if engagement["available"] and engagement["indicator"] >= 50:
            return "supported_by_engagement"

        if self._contains_any(lower, ("annual", "every year", "monthly")):
            return "recurring"

        return "possible"

    def _pattern_title(self, analysis, media_type):

        return (
            analysis.get("format", "Post format").replace("_", " ").title() +
            " / " +
            analysis.get("editorial_angle", media_type).replace("_", " ").title()
        )

    def _pattern_description(self, analysis):

        return (
            f"Uses a {analysis.get('hook_style', 'standard')} hook, "
            f"{analysis.get('educational_structure', 'single_message')} structure, "
            f"and {analysis.get('cta_type', 'no')} CTA."
        )

    def _limitations(self, items):

        if len(items) == 1:
            return "Single-record evidence; treat as inspiration, not a rule."

        if not any(item.get("engagement_available") for item in items):
            return "No engagement metrics supplied; performance is unverified."

        return "Comparable only within similar platform, source, and media context."

    def _preview_duplicates(self, records):

        counts = Counter(record.get("content_hash", "") for record in records)
        return sum(1 for value in counts.values() if value > 1)

    def _preview_confidence(self, source_type, mapping, warnings):

        required = ("department", "platform", "post_text")
        mapped = sum(1 for key in required if mapping.get(key))
        confidence = mapped * 30

        if source_type != "unsupported":
            confidence += 10

        confidence -= min(30, len(warnings) * 3)
        return max(0, min(100, confidence))

    def _parse_date(self, value):

        text = self._clean(value)

        if not text:
            return "", "Missing source date."

        formats = (
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%B %d, %Y",
            "%b %d, %Y"
        )

        for fmt in formats:
            try:
                parsed = datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
                return parsed.isoformat(), ""
            except Exception:
                pass

        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)
            return parsed.astimezone(timezone.utc).isoformat(), ""
        except Exception:
            return "", f"Ambiguous source date: {text}"

    def _hashtags(self, text):

        return list(dict.fromkeys(re.findall(r"#[A-Za-z0-9_]+", str(text or ""))))[:12]

    def _hash(self, text):

        return hashlib.sha256(str(text or "").encode("utf-8")).hexdigest()

    def _contains_any(self, lower, terms):

        return any(term in lower for term in terms)

    def _truthy(self, value):

        return str(value or "").strip().lower() in ("1", "yes", "true", "y", "reel", "short")

    def _token(self, value):

        value = str(value or "").strip().lower()
        value = re.sub(r"[^a-z0-9]+", "_", value)
        return value.strip("_") or "general"

    def _clean(self, value):

        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _to_int(self, value):

        try:
            return int(float(value or 0))
        except Exception:
            return 0

    def _to_float(self, value):

        try:
            return float(value or 0)
        except Exception:
            return 0
